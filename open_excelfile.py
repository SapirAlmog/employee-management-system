import openpyxl

theFile = openpyxl.load_workbook('Customers1.xlsx')
print(theFile.sheetnames)
currentSheet = theFile['customers 1']
print(currentSheet['B4'].value)

"""As you can see, this code prints all sheets by their names. It then selects the sheet that is named “customers 1”
 and saves it to a currentSheet variable. In the last line, the code prints the value that is located in the B4 
 position of the “customers 1” sheet.
"""

import openpyxl

theFile = openpyxl.load_workbook('Customers1.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))


for x in allSheetNames:
    print("Current sheet name is {}" .format(x))
    currentSheet = theFile[x]
    print(currentSheet['B4'].value)

"""Read the file
Get all sheet names
Loop through all sheets
In the last step, the code will print values that are located in B4 fields of each found sheet inside the workbook.

https://www.freecodecamp.org/news/how-to-create-read-update-and-search-through-excel-files-using-python-c70680d811d4/
"""
